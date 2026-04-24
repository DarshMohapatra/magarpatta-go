"""
Generate the Magarpatta Go test-cases workbook in the PauaPay reference format.

Reference: /mnt/c/Users/Darsh/Downloads/PauaPay_Indicative_Test_Cases_v2.xlsx

Columns (12): ✓ · Test Case ID · Priority · Test Scenario · Test Steps ·
             Expected Result · Actual Result · Status · Comments/Notes ·
             Tester Name · Date Tested · Completion Timestamp

Run: python3 docs/generate-test-cases.py
Output: docs/test-cases.xlsx
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Palette (matches PauaPay) ─────────────────────────────────────────────
HEADER_FILL = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
PRIORITY_STYLES = {
    "Critical": (PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"), Font(color="FFFFFF", bold=True)),
    "High":     (PatternFill(start_color="FFA07A", end_color="FFA07A", fill_type="solid"), Font(color="FFFFFF", bold=True)),
    "Medium":   (PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid"), Font(color="3A2E00", bold=True)),
    "Low":      (PatternFill(start_color="6BCF7F", end_color="6BCF7F", fill_type="solid"), Font(color="FFFFFF", bold=True)),
}
THIN = Side(style="thin", color="D7D7D7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(vertical="top", wrap_text=True)
CENTER = Alignment(vertical="center", horizontal="center", wrap_text=True)
LIGHT_FILL = PatternFill(start_color="F6F5FB", end_color="F6F5FB", fill_type="solid")

# ── Columns — must match PauaPay order/width ──────────────────────────────
COLUMNS = [
    ("✓",                    5),
    ("Test Case ID",         16),
    ("Priority",             12),
    ("Test Scenario",        40),
    ("Test Steps",           52),
    ("Expected Result",      44),
    ("Actual Result",        32),
    ("Status",               12),
    ("Comments/Notes",       34),
    ("Tester Name",          18),
    ("Date Tested",          15),
    ("Completion Timestamp", 22),
]

def steps(pre, *items):
    prefix = f"Pre: {pre}\n" if pre else ""
    return prefix + "\n".join(f"{i+1}. {s}" for i, s in enumerate(items))

# ── Test cases, keyed by sheet name ──────────────────────────────────────
# Row schema: (id, priority, scenario, test_steps, expected, comments)

SHEETS = {
    "Customer App": [
        ("TC-CUST-001", "Critical", "Phone OTP sign-in — happy path",
         steps("Firebase phone auth is live; the user is not yet signed in",
               "Open /",
               "Click 'Sign in'",
               "Enter a valid 10-digit Indian mobile",
               "Tap 'Get OTP'",
               "Enter the 6-digit code from SMS",
               "Tap 'Verify'"),
         "User redirected to home; navbar shows phone in account menu; mg_session cookie set.",
         "Auth / phone-OTP. Test data: any +91 mobile you control."),

        ("TC-CUST-002", "Critical", "Phone OTP sign-in — wrong OTP",
         steps("On the OTP screen after tapping Get OTP",
               "Enter six random digits",
               "Tap Verify"),
         "Inline error 'Invalid code. Try again.' Form stays on OTP step.",
         "Security. Test data: OTP=000000."),

        ("TC-CUST-003", "Medium", "OTP resend rate-limit",
         steps("OTP has been requested once",
               "Tap 'Resend OTP' three times within 10s"),
         "Resend button disabled with visible countdown until the rate-limit window expires.",
         "Auth UX."),

        ("TC-CUST-004", "High", "Delivery address — pick society / building / flat",
         steps("Signed in with no saved address",
               "Pick a society from the tower-select modal",
               "Pick a building",
               "Enter a flat number",
               "Save"),
         "Address saved; navbar chip shows 'Flat <n>, <building>'. Subsequent orders auto-use this address.",
         "Test data: Aspen / Aspen-2 / 402."),

        ("TC-CUST-005", "Medium", "Category rail renders",
         steps("Catalog has been seeded",
               "Open /"),
         "Category row shows Sweets & Snacks, Fresh Produce, Dairy, Pantry Staples, Bakery, Meat, Medicines, Beverages, Essentials.",
         "UI smoke."),

        ("TC-CUST-006", "Medium", "Restaurants grid shows seeded vendors",
         steps("Catalog + dashboards seeds ran",
               "Scroll to Restaurants / open /restaurants"),
         "All APPROVED + active vendors render: Kalika, Baker's, Destination Centre, Shraddha Meats, MG Pharmacy, Starbucks.",
         "UI."),

        ("TC-CUST-007", "High", "Cart — add single item",
         steps("Browsing menu",
               "Tap Add on Hot Jalebi (Kalika)",
               "Open cart drawer"),
         "Drawer header reads 'From Magarpatta Market'. Item line shows qty 1 at ₹120.",
         "Functional."),

        ("TC-CUST-008", "High", "Cart hub rule — same-hub multi-vendor allowed",
         steps("Kalika (Magarpatta Market) + Shraddha Meats (Magarpatta Market)",
               "Add Hot Jalebi from Kalika",
               "Scroll to Shraddha Meats",
               "Add Chicken Breast"),
         "Both items land in cart. Drawer header: 'From Magarpatta Market'. Sub-line: 'Kalika Sweets + Shraddha Meats · one hub, one trip'.",
         "Hub rule."),

        ("TC-CUST-009", "High", "Cart hub rule — cross-hub add blocked",
         steps("Cart already has a Kalika item",
               "Try to add any Starbucks (Seasons Mall) item"),
         "HubSwitchDialog opens: 'One trip, one hub' with 'Keep Magarpatta Market' / 'Clear & switch' options. Neither item is added automatically.",
         "Hub rule."),

        ("TC-CUST-010", "Medium", "Cart — increment / decrement / remove",
         steps("Cart has an item",
               "Open drawer, tap + twice, tap − once",
               "Tap 'remove' on any line"),
         "Qty goes 1→3→2. After Remove, line disappears; totals recompute; empty-state renders if cart is empty.",
         "Functional."),

        ("TC-CUST-011", "Critical", "Checkout — happy path (COD)",
         steps("Cart has items; signed in; address saved",
               "Open /checkout",
               "Confirm address step",
               "Choose Cash on Delivery",
               "Click Place order"),
         "Order row created: status=PLACED; fulfilmentMode set by vendor.supportsSelfDelivery; Order.hub snapshotted. Redirect to /orders/<id>.",
         "Core business flow."),

        ("TC-CUST-012", "High", "Checkout blocked without address",
         steps("Cart has items; signed in; no saved address",
               "Click Checkout"),
         "Checkout page shows 'Set a delivery address first' CTA; Place order button is disabled.",
         "Validation."),

        ("TC-CUST-013", "Medium", "Coupon — apply valid",
         steps("Active coupon seeded where minSubtotal ≤ cart subtotal",
               "Enter code in checkout",
               "Tap Apply"),
         "Discount row appears; total drops by (percent × subtotal) or flat amount.",
         "Test data: WELCOME50."),

        ("TC-CUST-014", "Medium", "Coupon — invalid rejected",
         steps("At checkout",
               "Enter an unknown code",
               "Tap Apply"),
         "Error 'Invalid coupon'. Total unchanged.",
         "Test data: NOPE."),

        ("TC-CUST-015", "High", "Order tracker — fresh order stays at PLACED (regression)",
         steps("Place a brand-new order and open /orders/<id>",
               "Wait 3 minutes without acting",
               "Inspect status chip + timeline"),
         "Status remains 'Order placed'. Timeline timestamps show '—' for all steps except PLACED. No demo auto-progression.",
         "Regression — demo fast-forward removed."),

        ("TC-CUST-016", "High", "Order tracker — real actor timestamps populate",
         steps("Open /orders/<id> on tab A; separate tabs for vendor + rider",
               "Vendor accepts",
               "Vendor marks ready (platform-rider flow)",
               "Rider claims",
               "Rider picks up",
               "Rider delivers with correct OTP"),
         "Within 3-second poll, each step moves to done and shows an IST timestamp. Final state: DELIVERED with every step green.",
         "End-to-end tracker sync."),

        ("TC-CUST-017", "High", "Order tracker — OTP shown only while inbound",
         steps("Order is ACCEPTED / PREPARING / PICKED_UP / OUT_FOR_DELIVERY",
               "Open /orders/<id>"),
         "4-digit OTP card visible, readable by the customer to hand over at drop.",
         "Security / UX."),

        ("TC-CUST-018", "Medium", "Order tracker — OTP hidden after delivery",
         steps("Same order is now DELIVERED",
               "Refresh /orders/<id>"),
         "OTP card is gone; 'Delivered' card remains.",
         "UX."),

        ("TC-CUST-019", "High", "Vendor-self tracker variant",
         steps("Baker's Basket (supportsSelfDelivery=true). Place an order from Baker's",
               "Sign in as Baker's vendor",
               "Click 'Head out now'",
               "Observe the customer tracker"),
         "SVG swaps bike icon for walking courier. 'Vendor-direct delivery' pill appears. Sub-line reads '<vendor> left at HH:MM'. No rider card.",
         "UI."),

        ("TC-CUST-020", "High", "Feedback — appears only after delivery",
         steps("Order status is DELIVERED",
               "Open /orders/<id>",
               "Scroll past the tracker"),
         "FeedbackForm renders with two star rows (food, delivery) + comment boxes. Does NOT render when status < DELIVERED.",
         "Phase-2 feature."),

        ("TC-CUST-021", "High", "Feedback — submit food + delivery rating",
         steps("On /orders/<id> (DELIVERED)",
               "Tap 4 stars under Food",
               "Tap 5 stars under Delivery",
               "Type a short comment in each",
               "Click Submit rating"),
         "POST /api/orders/<id>/feedback succeeds. OrderFeedback row created with foodRating=4, deliveryRating=5 linked to vendorId + riderPhone. Toast: 'Thanks! Feedback saved.'",
         "Writes to OrderFeedback table."),

        ("TC-CUST-022", "Medium", "Feedback — edit after submission",
         steps("An existing OrderFeedback row for this order",
               "Reload /orders/<id>",
               "Change food rating to 3 and edit comment",
               "Click Update rating"),
         "POST /api/orders/<id>/feedback upserts. Existing row now shows foodRating=3. Banner reads 'Already rated · edit if you like'.",
         "Idempotent upsert."),

        ("TC-CUST-023", "Medium", "Feedback — block without rating",
         steps("On feedback form with both ratings at 0",
               "Click Submit rating"),
         "Inline error 'Please rate at least one of food or delivery.' No API call.",
         "Validation."),

        ("TC-CUST-024", "Medium", "Orders list — reorder",
         steps("At least one past DELIVERED order",
               "Open /orders",
               "Click Reorder on a row"),
         "Cart clears, same items re-added, drawer opens automatically.",
         "Functional."),

        ("TC-CUST-025", "Critical", "API — /api/orders requires session",
         steps("Signed out",
               "curl https://<host>/api/orders"),
         "401 { ok:false, error:'Not signed in' }",
         "Security."),

        ("TC-CUST-026", "Critical", "API — cross-hub cart rejected",
         steps("POST /api/orders with productIds from two hubs",
               "Body mixes one Kalika (Magarpatta Market) + one Starbucks (Seasons Mall) productId"),
         "400 'Your cart has items from multiple hubs. Split into separate orders.' No order created.",
         "API hub rule."),
    ],

    "Vendor": [
        ("TC-VEN-001", "High", "Register — 3-step happy path → PENDING",
         steps("Unregistered phone",
               "Open /vendor/register",
               "Fill Shop step (name, type, hub, address, hours, owner, password)",
               "Next → KYC step (FSSAI, PAN, optional GSTIN / Drug licence)",
               "Next → Payouts step (bank, UPI)",
               "Submit"),
         "'Thanks, neighbour' success page. DB: Vendor row with approvalStatus=PENDING, active=false, ownerPasswordHash set.",
         "Password must be 6+ chars."),

        ("TC-VEN-002", "High", "Register — duplicate phone rejected",
         steps("Target phone is already a vendor owner",
               "Open /vendor/register",
               "Enter that phone",
               "Submit"),
         "409 'A vendor account with this phone already exists.' Form state preserved.",
         "Duplicate-phone uniqueness."),

        ("TC-VEN-003", "Medium", "Register — weak password blocked",
         steps("On /vendor/register step 1",
               "Enter a 3-character password",
               "Try to Continue"),
         "Continue button disabled; helper text mentions 6+ chars.",
         "Validation."),

        ("TC-VEN-004", "Critical", "Sign in — correct credentials",
         steps("Approved vendor exists",
               "Open /vendor/signin",
               "Enter 9000000001 / kalika123",
               "Submit"),
         "Redirect to /vendor. Overview shows 'Hello, <owner first name>'.",
         "Auth. Use ACCESS.md."),

        ("TC-VEN-005", "High", "Sign in — pending vendor blocked",
         steps("Demo Dosa House is PENDING",
               "Open /vendor/signin",
               "Enter 9000000003 / dosa123"),
         "Redirect lands on 'Under review' splash OR inline error. Vendor dashboard actions remain disabled.",
         "Approval gate."),

        ("TC-VEN-006", "Critical", "Sign in — wrong password",
         steps("Approved vendor",
               "Enter 9000000001 / wrong-password"),
         "401 'Wrong password.' Stays on signin.",
         "Security."),

        ("TC-VEN-007", "Medium", "Overview KPIs render",
         steps("Signed in as approved vendor",
               "Land on /vendor"),
         "Four stat cards render: Today's sales, Today's payout, New orders, In kitchen. Numbers match the DB for that vendor.",
         "UI."),

        ("TC-VEN-008", "Critical", "Menu — MRP rule for non-regulated (+₹1)",
         steps("On /vendor/menu",
               "Click + Add item",
               "Name=Test Samosa, Category=Sweets & Snacks, MRP=40, isRegulated=false, isVeg=true",
               "Save"),
         "POST /api/vendor/products returns priceInr=41 (MRP + ₹1 markup auto-applied). Row rendered under the correct category.",
         "Legal Metrology compliance."),

        ("TC-VEN-009", "Critical", "Menu — MRP rule for regulated (price = MRP)",
         steps("On /vendor/menu",
               "Add item with MRP=28, isRegulated=true",
               "Save"),
         "priceInr=28. Price field locked to MRP in the drawer UI.",
         "Legal Metrology compliance."),

        ("TC-VEN-010", "Medium", "Menu — toggle stock off",
         steps("An in-stock product exists",
               "Flip the In-stock checkbox off"),
         "Product disappears from the customer menu immediately. DB: inStock=false.",
         "Functional."),

        ("TC-VEN-011", "Medium", "Menu — edit price on existing item",
         steps("Non-regulated product exists",
               "Click Edit",
               "Change MRP to 50",
               "Save"),
         "mrpInr=50, priceInr=51 (MRP + ₹1 markup).",
         "Functional."),

        ("TC-VEN-012", "Medium", "Menu — soft-delete item",
         steps("Product exists",
               "Click Remove on the row",
               "Confirm"),
         "Row gone from vendor list. DB: inStock=false. Historical OrderItems still reference the product.",
         "Data integrity."),

        ("TC-VEN-013", "High", "Orders queue — new order appears on poll",
         steps("Signed in as Kalika",
               "Customer places an order from Kalika",
               "Wait up to 5s"),
         "Order card appears under 'New · tap to accept' with saffron accent. Fulfilment badge shows 'Rider pickup'.",
         "5s auto-refresh."),

        ("TC-VEN-014", "Critical", "Orders — Accept order",
         steps("A PLACED order exists for this vendor",
               "Click Accept on the card"),
         "Moves to 'Preparing'. DB: status=ACCEPTED, vendorAcceptedAt=now, acceptedAt=now.",
         "Core business flow."),

        ("TC-VEN-015", "High", "Orders — Reject with reason",
         steps("A PLACED order exists",
               "Click Reject",
               "Enter reason 'Out of stock'",
               "Confirm"),
         "Order moves to Today's history. DB: status=CANCELLED, cancelledAt=now, cancelReason='Out of stock'.",
         "Customer sees the reason."),

        ("TC-VEN-016", "High", "Orders — Platform-rider flow buttons",
         steps("Signed in as Kalika (supportsSelfDelivery=false)",
               "Accept an order",
               "Click 'Ready for pickup'"),
         "Card carries 'Rider pickup' badge. After Ready, it moves to 'Ready · waiting for pickup'. No 'Head out now' button appears.",
         "Conditional UI per fulfilmentMode."),

        ("TC-VEN-017", "High", "Orders · Self-delivery — head out",
         steps("Signed in as Baker's (supportsSelfDelivery=true)",
               "Accept order",
               "Click 'Mark ready (boxed)'",
               "Click 'Head out now →'"),
         "DB: pickedUpAt=now, status=OUT_FOR_DELIVERY. Card badge shows 'You deliver'.",
         "Vendor-self flow."),

        ("TC-VEN-018", "Critical", "Orders · Self-delivery — delivered via OTP",
         steps("Order is OUT_FOR_DELIVERY",
               "Click 'Mark delivered (OTP)'",
               "Enter the OTP visible on the customer's order page"),
         "DB: status=DELIVERED, deliveredAt=now. Card leaves the board.",
         "Delivery confirmation."),

        ("TC-VEN-019", "Critical", "Orders · Self-delivery — wrong OTP rejected",
         steps("OUT_FOR_DELIVERY order",
               "Mark delivered",
               "Enter 0000"),
         "400 'Wrong OTP. Ask the customer to check their order page.' Status unchanged.",
         "Security."),

        ("TC-VEN-020", "Medium", "Shop — edit description + save",
         steps("On /vendor/shop",
               "Change description",
               "Click Save changes"),
         "'Saved ✓' toast. Customer vendor card reflects new description after refresh.",
         "Functional."),

        ("TC-VEN-021", "High", "Shop — toggle self-delivery on",
         steps("On /vendor/shop",
               "Flip 'I deliver my own orders' on",
               "Save"),
         "DB: Vendor.supportsSelfDelivery=true. Next order placed from this vendor lands with fulfilmentMode=VENDOR_SELF.",
         "Admin has the same toggle."),

        ("TC-VEN-022", "Medium", "Shop — pause live",
         steps("Approved, active vendor",
               "Click 'Pause shop'"),
         "active=false. Customer listings hide the shop. In-flight orders continue unaffected.",
         "Functional."),

        ("TC-VEN-023", "Medium", "Payouts — 14-day breakdown renders",
         steps("A few delivered orders exist",
               "Open /vendor/payouts"),
         "Gross / Commission / Net tiles populated. Per-day bar chart renders. Recent delivered orders list shows last 20.",
         "UI."),

        ("TC-VEN-024", "High", "Feedback page — avg + distribution",
         steps("Customers have rated food on past delivered orders",
               "Sign in as that vendor",
               "Open /vendor/feedback"),
         "Avg food rating tile shows average across all ratings. 1–5 distribution bar chart renders proportionally. Recent ratings list with comments.",
         "Phase-2 vendor insights."),

        ("TC-VEN-025", "Medium", "Feedback page — empty state",
         steps("Vendor with no ratings yet",
               "Open /vendor/feedback"),
         "'No feedback yet. Ratings show up here as soon as delivered orders get rated.'",
         "UI."),

        ("TC-VEN-026", "Critical", "API — /api/vendor/orders requires vendor session",
         steps("Signed out or signed in as a customer",
               "curl /api/vendor/orders"),
         "401 'Not signed in'.",
         "Security."),

        ("TC-VEN-027", "Critical", "API — cross-vendor action blocked",
         steps("Signed in as Baker's",
               "POST /api/vendor/orders/<kalika-order-id>/accept"),
         "403 'Not your order.' No DB change.",
         "Security."),
    ],

    "Rider": [
        ("TC-RIDE-001", "High", "Register — apply to ride",
         steps("A phone not in RiderProfile",
               "Open /rider/register",
               "Fill name, phone, Aadhaar, DL, vehicle",
               "Submit"),
         "'Application in, neighbour' success page. DB: RiderProfile row with approvalStatus=PENDING.",
         "Self-serve signup."),

        ("TC-RIDE-002", "High", "Register — duplicate phone rejected",
         steps("An existing rider row (8888888801 — Akash)",
               "Open /rider/register",
               "Enter 8888888801",
               "Submit"),
         "409 'This phone is already a rider — sign in instead.'",
         "Uniqueness."),

        ("TC-RIDE-003", "Medium", "Register — missing fields",
         steps("On /rider/register",
               "Leave Full name blank",
               "Try submit"),
         "Submit button disabled until name + 10-digit phone are filled.",
         "Validation."),

        ("TC-RIDE-004", "Critical", "Sign in — approved rider",
         steps("Phone is in RiderProfile with approvalStatus=APPROVED",
               "Open /rider/signin",
               "Enter 8888888801",
               "Tap Start shift"),
         "Redirect to /rider dashboard. mg_rider_session cookie set.",
         "Auth."),

        ("TC-RIDE-005", "High", "Sign in — pending rider blocked",
         steps("A PENDING rider",
               "Enter their phone",
               "Tap Start shift"),
         "403 'Your application is still under review. We'll call you soon.'",
         "Approval gate."),

        ("TC-RIDE-006", "Medium", "Sign in — rejected rider",
         steps("A REJECTED rider",
               "Enter their phone"),
         "403 'Your application was not approved. Contact ops if you think this is wrong.'",
         "Approval gate."),

        ("TC-RIDE-007", "Medium", "Sign in — unknown phone",
         steps("Phone not in any RiderProfile row",
               "Enter a random phone",
               "Submit"),
         "401 'No rider application on this phone. Register first to apply.' with link to /rider/register.",
         "UX."),

        ("TC-RIDE-008", "Critical", "Available guardrail — 60-min stale cutoff",
         steps("Order placed at T−90 min and vendor-accepted at T−75 min; still unassigned",
               "Sign in as rider",
               "Check Available now"),
         "The stale order does NOT appear. Only orders with vendorAcceptedAt within the last 60 minutes show.",
         "Phase-2 rider guardrail."),

        ("TC-RIDE-009", "High", "Available guardrail — PLACED orders hidden",
         steps("A PLACED order exists (vendor has not accepted)",
               "Sign in as rider",
               "Observe Available now"),
         "The PLACED order does NOT appear. Only orders with status ∈ {ACCEPTED, PREPARING} show.",
         "Guardrail."),

        ("TC-RIDE-010", "High", "Available guardrail — vendor-self orders skipped",
         steps("Baker's Basket (supportsSelfDelivery=true) accepts an order",
               "Sign in as rider"),
         "The Baker's order never appears in the rider's Available list (fulfilmentMode=VENDOR_SELF filter).",
         "Guardrail."),

        ("TC-RIDE-011", "Critical", "Claim (accept) an order",
         steps("A vendor-accepted PLATFORM_RIDER order exists",
               "Click Accept on the card"),
         "Card moves to 'Your active deliveries'. DB: riderPhone, riderName, riderAssignedAt set.",
         "Core business flow."),

        ("TC-RIDE-012", "High", "Pickup — status transition",
         steps("An ACCEPTED order assigned to this rider",
               "Open its order page",
               "Tap 'Picked up'"),
         "DB: pickedUpAt=now, status=PICKED_UP. Customer tracker derives OUT_FOR_DELIVERY from the timestamp.",
         "Functional."),

        ("TC-RIDE-013", "Critical", "Delivery — OTP verified",
         steps("A picked-up order",
               "Ask customer for OTP",
               "Enter correct OTP",
               "Submit"),
         "DB: status=DELIVERED, deliveredAt=now. Today's earnings increment by ₹30.",
         "Core business flow."),

        ("TC-RIDE-014", "Critical", "Delivery — wrong OTP rejected",
         steps("A picked-up order",
               "Enter an incorrect 4-digit code"),
         "400 'Wrong OTP.' Status unchanged.",
         "Security."),

        ("TC-RIDE-015", "High", "Feedback page — avg + distribution",
         steps("Customers have rated rider's delivered drops",
               "Open /rider/feedback"),
         "Avg delivery rating tile shows average across ratings. Distribution chart + recent comments render.",
         "Phase-2 insights."),

        ("TC-RIDE-016", "Medium", "Sign out clears cookie",
         steps("Signed in as rider",
               "Click Sign out"),
         "Redirect to /rider/signin. Subsequent /api/rider/orders calls return 401.",
         "Auth."),
    ],

    "Admin": [
        ("TC-ADM-001", "Critical", "Sign in — valid",
         steps("Admin seeded (9999999999 / admin123)",
               "Open /admin/signin",
               "Enter credentials",
               "Submit"),
         "Redirect to /admin. mg_admin_session cookie set.",
         "Auth."),

        ("TC-ADM-002", "Critical", "Sign in — wrong password",
         steps("Admin seeded",
               "Enter 9999999999 / wrong"),
         "401 'Invalid credentials.' Stays on signin.",
         "Security."),

        ("TC-ADM-003", "Critical", "Protected page redirects when signed out",
         steps("No admin cookie",
               "Open /admin/vendors directly"),
         "Redirect to /admin/signin.",
         "Security."),

        ("TC-ADM-004", "Medium", "Overview KPI tiles",
         steps("Signed in as admin with data present",
               "Land on /admin"),
         "Four top tiles render (Pending vendors, Pending riders, Active orders, Today's GMV). Three mini-tiles underneath. Quick actions.",
         "UI."),

        ("TC-ADM-005", "High", "Vendors — Pending queue",
         steps("Seed applied",
               "Open /admin/vendors (default tab = Pending)"),
         "Demo Dosa House appears in Pending with owner + hub.",
         "Approval workflow."),

        ("TC-ADM-006", "Critical", "Vendors — approve + activate",
         steps("A PENDING vendor selected in the KYC drawer",
               "Click 'Approve + activate'"),
         "DB: approvalStatus=APPROVED, active=true, approvedAt stamped. Vendor moves to Approved tab.",
         "Core ops flow."),

        ("TC-ADM-007", "High", "Vendors — reject with note",
         steps("A PENDING vendor",
               "Type reviewer note",
               "Click Reject"),
         "DB: approvalStatus=REJECTED, approvalNote stored. Signin attempts show the note.",
         "Approval workflow."),

        ("TC-ADM-008", "High", "Vendors — suspend / lift",
         steps("An APPROVED vendor",
               "Click Suspend + note",
               "Later click 'Lift suspension' on the suspended row"),
         "Suspend: SUSPENDED + active=false. Lift: back to APPROVED + active=true.",
         "Ops control."),

        ("TC-ADM-009", "High", "Vendors — fulfilment toggle live",
         steps("Approved vendor open in drawer",
               "Flip 'Supports self-delivery'"),
         "POST /api/admin/vendors/<id>/fulfilment succeeds. DB field flipped. Subsequent new orders from this vendor use the new fulfilmentMode.",
         "Phase-2 control."),

        ("TC-ADM-010", "High", "Riders — Pending queue shows Kiran J.",
         steps("Seed applied",
               "Open /admin/riders"),
         "Kiran J. visible in Pending tab.",
         ""),

        ("TC-ADM-011", "High", "Riders — onboard rider (admin)",
         steps("Signed in as admin",
               "Click '+ Onboard rider'",
               "Fill form",
               "Submit"),
         "New RiderProfile PENDING row created from the form values.",
         ""),

        ("TC-ADM-012", "Critical", "Riders — approve",
         steps("Pending rider",
               "Click Approve"),
         "DB: approvalStatus=APPROVED, approvedAt stamped. Rider can sign in at /rider/signin.",
         "Core ops flow."),

        ("TC-ADM-013", "High", "Riders — Performance tab",
         steps("Seed applied; some delivered orders exist",
               "Open /admin/riders",
               "Switch to Performance tab"),
         "Table shows per-rider: today drops, 30-day drops + earnings, avg pickup→drop minutes, avg rating + rating count. Sorted by 30-day drops desc.",
         "Phase-2 ops insight."),

        ("TC-ADM-014", "Critical", "Orders — Active scope excludes terminal states",
         steps("Mix of active, delivered, and cancelled orders",
               "Open /admin/orders (default scope=active)"),
         "Only orders with status ∈ {PLACED, ACCEPTED, PREPARING, PICKED_UP, OUT_FOR_DELIVERY} render. DELIVERED and CANCELLED never show here.",
         "Phase-2 fix."),

        ("TC-ADM-015", "High", "Orders — Delivered tab",
         steps("Delivered orders exist",
               "Switch tab to Delivered"),
         "Only DELIVERED orders show. Each row shows deliveredAt stamp. No Reassign/Cancel buttons.",
         "Phase-2 tab."),

        ("TC-ADM-016", "High", "Orders — Cancelled tab",
         steps("Cancelled orders exist (by vendor or admin)",
               "Switch tab to Cancelled"),
         "Only CANCELLED orders show with cancelledAt + 'Reason: …' line.",
         "Phase-2 tab."),

        ("TC-ADM-017", "High", "Orders — reassign rider",
         steps("An active order with an assigned rider",
               "Click Reassign rider",
               "Pick a different approved rider",
               "Save"),
         "DB: riderPhone / riderName updated. The new rider sees the order in their Active list.",
         "Ops control."),

        ("TC-ADM-018", "High", "Orders — cancel with refund note",
         steps("An active order",
               "Click 'Cancel + refund'",
               "Enter reason"),
         "DB: status=CANCELLED, cancelledAt + cancelReason stored. Moves to Cancelled tab.",
         "Ops control."),

        ("TC-ADM-019", "Medium", "Customers — search by phone",
         steps("Customer with phone 9876543210 exists",
               "Open /admin/customers",
               "Type the phone"),
         "Row filters to matching customer. Lifetime total + order count visible.",
         "UI."),

        ("TC-ADM-020", "Medium", "Finance — breakdown by day + vendor",
         steps("Some delivered orders in last 14 days",
               "Open /admin/finance"),
         "GMV / Gross merch / Delivery fee / Platform commission tiles populated. By-day bar chart. By-vendor ranking with commission.",
         "Ops insight."),

        ("TC-ADM-021", "Critical", "API — admin endpoints need admin cookie",
         steps("Signed in as a vendor (not admin)",
               "curl /api/admin/vendors"),
         "401 'Not signed in.'",
         "Security."),
    ],

    "Integration": [
        ("TC-INT-001", "Critical", "E2E — platform-rider flow",
         steps("Kalika Sweets is platform-rider; customer signed in; rider signed in",
               "Customer orders from Kalika",
               "Kalika vendor accepts",
               "Kalika marks ready",
               "Rider claims",
               "Rider picks up",
               "Rider delivers with correct OTP",
               "Customer refreshes /orders/<id>"),
         "Every step timestamp appears in DB in order. Tracker ends at DELIVERED with IST stamps under every step. Rider earnings +₹30.",
         "End-to-end."),

        ("TC-INT-002", "Critical", "E2E — vendor-self-delivery flow",
         steps("Baker's Basket (supportsSelfDelivery=true)",
               "Customer orders from Baker's",
               "Baker's accepts",
               "Marks ready",
               "Heads out now",
               "Delivers with correct OTP"),
         "Order never appears in the rider queue. Customer sees courier glyph (not bike). DB: fulfilmentMode=VENDOR_SELF, no riderPhone.",
         "End-to-end vendor-self."),

        ("TC-INT-003", "High", "Mixed fulfilment — any non-self vendor forces PLATFORM_RIDER",
         steps("Two same-hub vendors where one is supportsSelfDelivery=true and one is false",
               "Add items from both vendors in the same hub",
               "Place order"),
         "Order.fulfilmentMode resolves to PLATFORM_RIDER (because any non-self vendor requires a rider).",
         "Business rule."),

        ("TC-INT-004", "Medium", "Cart hub lock persists across sessions",
         steps("Cart has Kalika item; close then reopen browser as same user",
               "Try to add a Seasons Mall item"),
         "Cart restored from localStorage. Cross-hub block still triggers.",
         "Persistence."),

        ("TC-INT-005", "Critical", "Regression — no demo fast-forward when idle",
         steps("Place order at T=0 and leave it alone",
               "Wait 10 minutes on /orders/<id>",
               "Inspect DB"),
         "Status remains PLACED. Only placedAt is stamped. No demo auto-progression on view or on poll.",
         "Phase-2 fix."),

        ("TC-INT-006", "Critical", "API — server rejects cross-hub cart",
         steps("Build a POST body to /api/orders with items from two hubs",
               "POST"),
         "400 'Your cart has items from multiple hubs.' No Order row created.",
         "API hub rule."),

        ("TC-INT-007", "Critical", "MRP rule — non-regulated auto +₹1",
         steps("POST /api/vendor/products as approved vendor",
               "Body: isRegulated=false, mrpInr=40, priceInr missing"),
         "Created row: priceInr=41.",
         "Legal Metrology."),

        ("TC-INT-008", "Critical", "MRP rule — regulated locked to MRP",
         steps("POST /api/vendor/products",
               "Body: isRegulated=true, mrpInr=28, priceInr=30"),
         "Server overrides; stored priceInr=28.",
         "Legal Metrology."),

        ("TC-INT-009", "Critical", "Auth boundary — vendor can't call rider APIs",
         steps("Signed in as vendor only",
               "POST /api/rider/orders/<id>/accept"),
         "401 from rider-session. No DB change.",
         "Security."),

        ("TC-INT-010", "Critical", "Auth boundary — rider can't read vendor dashboard",
         steps("Signed in as rider only",
               "GET /api/vendor/orders"),
         "401.",
         "Security."),

        ("TC-INT-011", "High", "Feedback E2E — customer ↔ vendor ↔ rider",
         steps("Complete a platform-rider order end-to-end",
               "Customer rates food 4 + delivery 5 with comments",
               "Check /vendor/feedback (vendor side)",
               "Check /rider/feedback (rider side)"),
         "Vendor sees 4-star food rating + comment. Rider sees 5-star delivery rating + comment. Admin Performance tab shows avg rating for the rider.",
         "Phase-2 E2E."),

        ("TC-INT-012", "Medium", "Seed — idempotent",
         steps("Fresh clone of repo",
               "Run pnpm db:seed:dashboards twice"),
         "No duplicate rows. Counts stable across runs. Log output identical on second run.",
         "DevOps."),

        ("TC-INT-013", "High", "Deploy smoke — prod URL responds",
         steps("After a green Vercel deploy",
               "curl -I https://web-eta-ebon-80.vercel.app/"),
         "HTTP 200.",
         "Smoke."),
    ],
}

# ── Build workbook ────────────────────────────────────────────────────────

wb = Workbook()
wb.remove(wb.active)

# ─── Summary & Instructions sheet ────────────────────────────────────────
summary = wb.create_sheet("Summary & Instructions")
summary["A1"] = "Magarpatta Go — Indicative Test Cases"
summary["A1"].font = Font(bold=True, size=18, color="1E293B")
summary["A2"] = "Based on the Phase-2 build (vendor + admin dashboards, hub-locked cart, vendor-self-delivery, feedback system, rider guardrail)"
summary["A2"].font = Font(italic=True, color="475569")

instructions = [
    ("A4", "Instructions:", Font(bold=True, size=12)),
    ("A5", "1. Test cases aligned with Phase-2 feature set + existing Phase-1 customer flow.", None),
    ("A6", "2. Each category covers a specific role: Customer, Vendor, Rider, Admin, plus cross-role Integration.", None),
    ("A7", "3. Priority levels: Critical (security / compliance / data integrity) > High > Medium > Low.", None),
    ("A8", "4. Tick the ✓ column when a case is run. Enter Pass / Fail / Blocked / N/A in Status.", None),
    ("A9", "5. Record who tested and when. Keep notes in the Comments column.", None),
    ("A10", "6. Seed credentials live in docs/ACCESS.md — keep that open while testing.", None),
]
for cell, val, font in instructions:
    summary[cell] = val
    if font: summary[cell].font = font

summary["A12"] = "Phase-2 scope covered:"
summary["A12"].font = Font(bold=True, size=12)
scope_items = [
    "• Vendor dashboard with approval gate + self-delivery toggle",
    "• Admin dashboard (approvals, orders, customers, finance, rider performance)",
    "• Rider dashboard with strict 60-min 'available now' guardrail",
    "• Customer order tracker driven by real vendor + rider timestamps (no demo auto-progression)",
    "• Hub-locked cart — multi-vendor OK within one hub, cross-hub blocked",
    "• Two delivery modes: PLATFORM_RIDER and VENDOR_SELF (OTP on drop in both)",
    "• Customer feedback system (food + delivery stars, comments)",
    "• Vendor + rider read-back of customer feedback",
    "• Legal Metrology: regulated goods at MRP; non-regulated +₹1 markup",
    "• Admin orders Active / Delivered / Cancelled tabs with counts",
]
for i, item in enumerate(scope_items, start=13):
    summary[f"A{i}"] = item

# Summary table
start = 13 + len(scope_items) + 2
summary[f"A{start}"] = "Test Case Summary by Category:"
summary[f"A{start}"].font = Font(bold=True, size=12)

header_row = start + 1
summary_headers = ["Category", "Total", "Critical", "High", "Medium", "Low"]
for c, v in enumerate(summary_headers, start=1):
    cell = summary.cell(row=header_row, column=c, value=v)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = CENTER
    cell.border = BORDER

totals = {"Critical": 0, "High": 0, "Medium": 0, "Low": 0}
row_i = header_row + 1
for sheet_name, rows in SHEETS.items():
    c_ct = sum(1 for r in rows if r[1] == "Critical")
    h_ct = sum(1 for r in rows if r[1] == "High")
    m_ct = sum(1 for r in rows if r[1] == "Medium")
    l_ct = sum(1 for r in rows if r[1] == "Low")
    totals["Critical"] += c_ct; totals["High"] += h_ct; totals["Medium"] += m_ct; totals["Low"] += l_ct
    for c, v in enumerate([sheet_name, len(rows), c_ct, h_ct, m_ct, l_ct], start=1):
        cell = summary.cell(row=row_i, column=c, value=v)
        cell.alignment = Alignment(horizontal="center" if c > 1 else "left", vertical="center")
        cell.border = BORDER
        if row_i % 2 == 0: cell.fill = LIGHT_FILL
    row_i += 1
# Total row
grand_total = sum(len(r) for r in SHEETS.values())
for c, v in enumerate(["TOTAL", grand_total, totals["Critical"], totals["High"], totals["Medium"], totals["Low"]], start=1):
    cell = summary.cell(row=row_i, column=c, value=v)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center" if c > 1 else "left", vertical="center")
    cell.border = BORDER
    cell.fill = PatternFill(start_color="E9E5FA", end_color="E9E5FA", fill_type="solid")

# Project info
proj_start = row_i + 3
summary[f"A{proj_start}"] = "Project Information:"
summary[f"A{proj_start}"].font = Font(bold=True, size=12)
proj_rows = [
    ("Project:",       "Magarpatta Go — hyper-local delivery platform"),
    ("Phase:",         "Phase 2 MVP (vendor + admin + feedback)"),
    ("Platform:",      "Web (Next.js 15 App Router · Vercel)"),
    ("Database:",      "Postgres (Neon) via Prisma"),
    ("Deploy URL:",    "https://web-eta-ebon-80.vercel.app"),
    ("Credentials:",   "docs/ACCESS.md"),
    ("Document:",      "test-cases.xlsx · v2 (PauaPay format)"),
    ("Created:",       "2026-04-24"),
]
for i, (k, v) in enumerate(proj_rows, start=proj_start + 1):
    summary.cell(row=i, column=1, value=k).font = Font(bold=True)
    summary.cell(row=i, column=2, value=v)

for col, w in zip("ABCDEF", [34, 14, 12, 12, 12, 12]):
    summary.column_dimensions[col].width = w

# ─── Per-category data sheets ─────────────────────────────────────────────
for sheet_name, rows in SHEETS.items():
    ws = wb.create_sheet(sheet_name)

    # Header row
    for c_idx, (label, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=c_idx, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(c_idx)].width = width
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    # Data rows
    for r_idx, (tc_id, priority, scenario, test_steps, expected, comments) in enumerate(rows, start=2):
        # Column values: ✓, id, priority, scenario, steps, expected, actual, status, comments, tester, date, ts
        values = ["", tc_id, priority, scenario, test_steps, expected, "", "", comments, "", "", ""]
        for c_idx, v in enumerate(values, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=v)
            cell.alignment = WRAP
            cell.border = BORDER
            if r_idx % 2 == 0 and c_idx != 3:  # don't overwrite priority fill
                cell.fill = LIGHT_FILL
        # Priority styling
        pri_fill, pri_font = PRIORITY_STYLES[priority]
        pri_cell = ws.cell(row=r_idx, column=3)
        pri_cell.fill = pri_fill
        pri_cell.font = pri_font
        pri_cell.alignment = CENTER
        # ✓ centered
        ws.cell(row=r_idx, column=1).alignment = CENTER
        # Row height based on longest field
        longest = max(len(test_steps or ""), len(expected or ""))
        ws.row_dimensions[r_idx].height = max(52, min(200, 20 + longest // 8))

wb.save("/mnt/c/projects/magarpatta-delivery/docs/test-cases.xlsx")
print(f"Wrote docs/test-cases.xlsx · {grand_total} cases across {len(SHEETS)} sheets.")
print(f"  Critical: {totals['Critical']}  High: {totals['High']}  Medium: {totals['Medium']}  Low: {totals['Low']}")
